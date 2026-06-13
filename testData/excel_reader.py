import openpyxl

#SheetName
def get_excel_data(AddJobs):

    workbook = openpyxl.load_workbook(
        "testData/testdata.xlsx"
    )

    # SheetName
    sheet = workbook[AddJobs]

    total_rows = sheet.max_row
    total_cols = sheet.max_column

    data_list = []

    for row in range(2, total_rows + 1):

        row_data = []

        for col in range(1, total_cols + 1):

            row_data.append(
                sheet.cell(row=row, column=col).value
            )

        data_list.append(row_data)

    workbook.close()

    return data_list